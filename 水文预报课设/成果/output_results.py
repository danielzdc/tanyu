"""
完整成果输出: 日模型+次洪模型 各模块Excel表格 + 流量过程线图
"""
import sys, os, numpy as np
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from data_reader import read_xaj_data, extract_flood_data
from main_daily import run_daily_model_multi
from main_flood import run_flood_model_multi
from evaluation import evaluate_daily, evaluate_flood, calc_total_runoff
from sub_basin import run_sub_basin, muskingum_segment_routing

# ========== 配置 ==========
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter

# 中文字体
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# 路径
BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, '..', '呈村流域资料（含场次洪水）-0613更新场次洪水实测流量.xls')
OUT_DIR = os.path.join(BASE, 'output')
os.makedirs(OUT_DIR, exist_ok=True)

F = 290.0  # 总面积

# 参数
DAILY_PARAMS = {
    'K': 1.02, 'WUM': 20, 'WLM': 60, 'WDM': 40, 'C': 0.18,
    'WM': 120, 'b': 0.3, 'IM': 0.01,
    'SM': 22, 'EX': 1.5, 'KG': 0.15, 'KI': 0.55,
    'CI': 0.85, 'CG': 0.985, 'CS': 0.25, 'L': 0
}
FLOOD_PARAMS = {
    'K': 1.02, 'WUM': 20, 'WLM': 60, 'WDM': 40, 'C': 0.18,
    'WM': 120, 'b': 0.3, 'IM': 0.01,
    'SM': 36, 'EX': 1.5, 'KG': 0.030, 'KI': 0.110,
    'CI': 0.86, 'CG': 0.985, 'CS': 0.75, 'L': 1, 'X': 0.300
}

# ========== 加载数据 ==========
print("加载数据...")
daily_data, hourly_data, subs, events, param_ref = read_xaj_data(EXCEL)
sub_areas = subs['areas']
sub_nseg_flood = subs['m_routing_flood']

# 日E0映射
daily_e0_map = {}
for i, dt in enumerate(daily_data['dates']):
    daily_e0_map[dt.date()] = daily_data['E0'][i]
default_e0 = daily_data['E0'].mean()

# 预热状态
def warmup_daily(target_date, n_warmup=180):
    ti = None
    for i, dt in enumerate(daily_data['dates']):
        if dt.date() == target_date.date(): ti = i; break
    if ti is None:
        return [{'WU': DAILY_PARAMS['WUM']/3, 'WL': DAILY_PARAMS['WLM'], 'WD': DAILY_PARAMS['WDM'],
                 'S': 0, 'FR': 0, 'QI': 0, 'QG': 0} for _ in range(10)]
    we = min(n_warmup, ti)
    st = [{'WU': DAILY_PARAMS['WUM']/3, 'WL': DAILY_PARAMS['WLM'], 'WD': DAILY_PARAMS['WDM'],
           'S': 0, 'FR': 0, 'QI': 0, 'QG': 0} for _ in range(10)]
    if we > 0:
        for i in range(10):
            r = run_sub_basin(daily_data['P_stations'][:we, i], daily_data['E0'][:we],
                              sub_areas[i], 24, DAILY_PARAMS, st[i])
            st[i] = {'WU': r['WU_end'], 'WL': r['WL_end'], 'WD': r['WD_end'],
                     'S': r['S_end'], 'FR': r['FR_end'], 'QI': r['QI_end'], 'QG': r['QG_end']}
    if ti > we:
        for i in range(10):
            r = run_sub_basin(daily_data['P_stations'][we:ti, i], daily_data['E0'][we:ti],
                              sub_areas[i], 24, DAILY_PARAMS, st[i])
            st[i] = {'WU': r['WU_end'], 'WL': r['WL_end'], 'WD': r['WD_end'],
                     'S': r['S_end'], 'FR': r['FR_end'], 'QI': r['QI_end'], 'QG': r['QG_end']}
    return st

# 逐时E0
def make_hourly_E0(flood_start, n_h):
    FULL_DAY = 12.0
    h = np.zeros(n_h)
    for i in range(n_h):
        dt = flood_start + timedelta(hours=i)
        day_e0 = daily_e0_map.get(dt.date(), default_e0)
        raw = 0.5 + 0.45 * np.sin(2 * np.pi * (dt.hour - 8) / 24)
        h[i] = raw * day_e0 / FULL_DAY
    return h

# ========== Excel 工具 ==========
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=10)
DATA_FONT = Font(name='Microsoft YaHei', size=9)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_sheet(ws, headers, data, col_widths=None):
    """写表头+数据到worksheet"""
    for j, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=j, value=h)
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER
    for i, row in enumerate(data, 2):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val if not isinstance(val, float) or np.isfinite(val) else 0)
            cell.font = DATA_FONT; cell.border = THIN_BORDER
            if isinstance(val, float):
                cell.number_format = '0.000' if abs(val) < 10 else '0.00'
    if col_widths:
        for j, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(j)].width = w

# ========== 日模型 ==========
print("\n" + "="*60)
print("日模型计算...")

init_daily = [{'WU': DAILY_PARAMS['WUM']/3, 'WL': DAILY_PARAMS['WLM'], 'WD': DAILY_PARAMS['WDM'],
               'S': 0, 'FR': 0, 'QI': 0, 'QG': 0} for _ in range(10)]

# 选一个代表性子流域(面积最大的)做详细输出
rep_idx = int(np.argmax(sub_areas))
print(f"代表性子流域: sub_{rep_idx} (面积={sub_areas[rep_idx]:.1f}km2)")

# 跑每个子流域（详细模式）
daily_sub_results = []
for i in range(10):
    r = run_sub_basin(daily_data['P_stations'][:, i], daily_data['E0'],
                      sub_areas[i], 24, DAILY_PARAMS, init_daily[i], return_detailed=True)
    daily_sub_results.append(r)

# 总流量 = 各子流域直和 (CS+L 河网汇流)
Q_daily_sim = np.sum([daily_sub_results[i]['Q'] for i in range(10)], axis=0)
Q_daily_obs = daily_data['Q_obs']
dates_daily = daily_data['dates']
P_areal = np.sum([daily_data['P_stations'][:, i] * sub_areas[i] for i in range(10)], axis=0) / F

# 逐年统计
years_all = sorted(set(d.date().year for d in dates_daily))
daily_yearly = []
for yr in years_all:
    m = np.array([d.year == yr for d in dates_daily])
    qs, qo = Q_daily_sim[m], Q_daily_obs[m]
    Rs = qs.sum() * 24 * 3.6 / F
    Ro = qo.sum() * 24 * 3.6 / F
    re = abs(Rs - Ro) / max(Ro, 0.01) * 100
    w = qo**2; bo = np.sum(np.abs(qs - qo) * w) / max(np.sum(qo * w), 0.01)
    ok = "OK" if re <= 20 else "--"
    daily_yearly.append((yr, '率定' if yr <= 1994 else '检验', bo, re, Rs, Ro, ok))

# ========== 日模型 Excel ==========
print("输出日模型Excel...")
wb_daily = Workbook()
wb_daily.remove(wb_daily.active)

# 公用日期序列
date_strs = [d.strftime('%Y-%m-%d') for d in dates_daily]

# Sheet 1: 逐年结果
ws = wb_daily.create_sheet("逐年结果")
write_sheet(ws, ['年份', '类型', 'BOw', 'R_err(%)', 'R_sim(mm)', 'R_obs(mm)', '达标'],
            daily_yearly, [8,8,10,12,14,14,8])

# Sheet 2-6: 代表性子流域详细过程 (选某一年展示)
show_yr = 1992
show_mask = np.array([d.year == show_yr for d in dates_daily])
show_dates = [dates_daily[i] for i in range(len(dates_daily)) if show_mask[i]]
rep = daily_sub_results[rep_idx]

# 蒸散发
ws = wb_daily.create_sheet(f"蒸散发_{show_yr}")
hdr = ['日期', 'P(mm)', 'E0(mm)', 'EU(mm)', 'EL(mm)', 'ED(mm)', 'E(mm)', 'WU(mm)', 'WL(mm)', 'WD(mm)']
rows = []
for i in range(len(dates_daily)):
    if not show_mask[i]: continue
    rows.append([date_strs[i],
        round(daily_data['P_stations'][i, rep_idx], 2),
        round(daily_data['E0'][i], 2),
        round(rep['EU'][i], 3), round(rep['EL'][i], 3), round(rep['ED'][i], 3),
        round(rep['E'][i], 3),
        round(rep['WU'][i], 2), round(rep['WL'][i], 2), round(rep['WD'][i], 2)])
write_sheet(ws, hdr, rows, [12,8,8,10,10,10,10,10,10,10])

# 产流
ws = wb_daily.create_sheet(f"产流_{show_yr}")
hdr = ['日期', 'PE(mm)', 'W(mm)', 'a(mm)', 'R(mm)', 'FR']
rows = []
for i in range(len(dates_daily)):
    if not show_mask[i]: continue
    rows.append([date_strs[i],
        round(rep['PE'][i], 2), round(rep['WU'][i] + rep['WL'][i] + rep['WD'][i], 1),
        round(rep['a'][i], 2), round(rep['R'][i], 2), round(rep['FR_series'][i], 3)])
write_sheet(ws, hdr, rows, [12,10,10,10,10,10])

# 三水源
ws = wb_daily.create_sheet(f"三水源_{show_yr}")
hdr = ['日期', 'R(mm)', 'S(mm)', 'FR', 'RS(mm)', 'RI(mm)', 'RG(mm)']
rows = []
for i in range(len(dates_daily)):
    if not show_mask[i]: continue
    rows.append([date_strs[i],
        round(rep['R'][i], 2), round(rep['S'][i], 3),
        round(rep['FR_series'][i], 3),
        round(rep['RS'][i], 3), round(rep['RI'][i], 3), round(rep['RG'][i], 3)])
write_sheet(ws, hdr, rows, [12,10,10,10,10,10,10])

# 坡地汇流
ws = wb_daily.create_sheet(f"坡地汇流_{show_yr}")
hdr = ['日期', 'QS(m3/s)', 'QI(m3/s)', 'QG(m3/s)', 'QT(m3/s)']
rows = []
for i in range(len(dates_daily)):
    if not show_mask[i]: continue
    rows.append([date_strs[i],
        round(rep['QS'][i], 3), round(rep['QI_series'][i], 3),
        round(rep['QG_series'][i], 3), round(rep['QT'][i], 3)])
write_sheet(ws, hdr, rows, [12,12,12,12,12])

# 河网汇流
ws = wb_daily.create_sheet(f"河网汇流_{show_yr}")
hdr = ['日期', 'QT_lag(m3/s)', 'Q_sim(m3/s)']
rows = []
for i in range(len(dates_daily)):
    if not show_mask[i]: continue
    rows.append([date_strs[i],
        round(rep['QT_lagged'][i], 3), round(rep['Q'][i], 3)])
write_sheet(ws, hdr, rows, [12,14,14])

# 日流量过程(全时期)
ws = wb_daily.create_sheet("日流量过程")
hdr = ['日期', 'P_areal(mm)', 'Q_obs(m3/s)', 'Q_sim(m3/s)', 'error(m3/s)', 'R_err_daily(%)']
rows = []
for i in range(len(dates_daily)):
    qo = Q_daily_obs[i]; qs = Q_daily_sim[i]
    err = qs - qo
    re_day = abs(err) / max(qo, 0.01) * 100
    rows.append([date_strs[i],
        round(P_areal[i], 1), round(qo, 2), round(qs, 2),
        round(err, 3), round(re_day, 2)])
write_sheet(ws, hdr, rows, [12,10,12,12,12,12])

wb_daily.save(os.path.join(OUT_DIR, '日模型详细结果_v2.xlsx'))
print(f"  日模型Excel已保存: {os.path.join(OUT_DIR, '日模型详细结果.xlsx')}")

# ========== 次洪模型 ==========
print("\n" + "="*60)
print("次洪模型计算...")

flood_all = []
flood_results = []

for fe_idx, fe in enumerate(events['flood']):
    fd = extract_flood_data(hourly_data, fe)
    if fd is None or fd['n'] < 10: continue

    e0_h = make_hourly_E0(fe['start'], fd['n'])
    states = warmup_daily(fe['start'])
    is_rate = fe['start'].year <= 1994

    sub_results = []
    for i in range(10):
        r = run_sub_basin(fd['P_stations'][:, i], e0_h, sub_areas[i], 1.0,
                          FLOOD_PARAMS, states[i], return_detailed=True)
        # 马斯京根
        n_seg = int(sub_nseg_flood[i])
        if n_seg > 0:
            Q_routed = muskingum_segment_routing(r['Q'], n_seg, FLOOD_PARAMS['X'], 1.0)
        else:
            Q_routed = r['Q']
        sub_results.append({'result': r, 'Q_routed': Q_routed, 'n_seg': n_seg})

    # 流域出口总流量
    Q_total = np.sum([s['Q_routed'] for s in sub_results], axis=0)
    Q_obs = fd['Q_obs']

    Rs = Q_total.sum() * 3.6 / F
    Ro = calc_total_runoff(Q_obs, F, 1.0)
    ev = evaluate_flood(Q_total, Q_obs, Rs, Ro)
    w = Q_obs**2; bo = np.sum(np.abs(Q_total - Q_obs) * w) / max(np.sum(Q_obs * w), 0.01)
    ok = "OK" if ev['R_err'] <= 20 and ev['Qp_err'] <= 20 else ""

    flood_results.append({
        'no': fe_idx + 1, 'date': fe['start'], 'is_rate': is_rate,
        'BOw': bo, 'R_err': ev['R_err'], 'Qp_err': ev['Qp_err'],
        'Tp_err': ev['Tp_err'], 'DC': ev['DC'], 'ok': ok,
        'Q_sim': Q_total, 'Q_obs': Q_obs, 'n': fd['n'],
        'E0_h': e0_h, 'sub_results': sub_results, 'fd': fd
    })
    tag = "率" if is_rate else "检"
    print(f"  #{fe_idx+1} {tag} {fe['start'].strftime('%Y-%m-%d')}: "
          f"BOw={bo:.4f} R={ev['R_err']:.1f}% Qp={ev['Qp_err']:.1f}% {ok}")

# ========== 次洪模型 Excel ==========
print("\n输出次洪模型Excel...")
wb_flood = Workbook()
wb_flood.remove(wb_flood.active)

# 场次汇总
ws = wb_flood.create_sheet("场次洪水汇总")
hdr_sum = ['编号', '类型', '日期', 'BOw', 'R_err(%)', 'Qp_err(%)', 'Tp_err(h)', 'DC', '达标']
rows_sum = []
for fr in flood_results:
    rows_sum.append([fr['no'], '率定' if fr['is_rate'] else '检验',
        fr['date'].strftime('%Y-%m-%d'),
        round(fr['BOw'], 4), round(fr['R_err'], 1), round(fr['Qp_err'], 1),
        round(fr['Tp_err'], 0), round(fr['DC'], 4), fr['ok']])
write_sheet(ws, hdr_sum, rows_sum, [6,8,14,10,10,10,10,10,6])

# 每场洪水详细
for fr in flood_results:
    tag = f"#{fr['no']}_{fr['date'].strftime('%Y%m%d')}"
    fd = fr['fd']
    rep_sub = fr['sub_results'][int(np.argmax(sub_areas))]
    rep_r = rep_sub['result']
    rep_nseg = rep_sub['n_seg']

    date_strs_h = [(fr['date'] + timedelta(hours=i)).strftime('%m-%d %Hh') for i in range(fr['n'])]
    P_areal_h = np.sum([fd['P_stations'][:, i] * sub_areas[i] for i in range(10)], axis=0) / F

    # 蒸发
    ws = wb_flood.create_sheet(f"蒸发_{tag}"[:31])
    hdr = ['时间', 'P(mm)', 'E0(mm)', 'EU(mm)', 'EL(mm)', 'ED(mm)', 'E(mm)', 'WU(mm)', 'WL(mm)', 'WD(mm)']
    rows = []
    for i in range(fr['n']):
        rows.append([date_strs_h[i],
            round(fd['P_stations'][i, int(np.argmax(sub_areas))], 2),
            round(fr['E0_h'][i], 3),
            round(rep_r['EU'][i], 3), round(rep_r['EL'][i], 3), round(rep_r['ED'][i], 3),
            round(rep_r['E'][i], 3),
            round(rep_r['WU'][i], 2), round(rep_r['WL'][i], 2), round(rep_r['WD'][i], 2)])
    write_sheet(ws, hdr, rows, [14,8,10,10,10,10,10,10,10,10])

    # 产流
    ws = wb_flood.create_sheet(f"产流_{tag}"[:31])
    hdr = ['时间', 'PE(mm)', 'W(mm)', 'a(mm)', 'R(mm)', 'FR']
    rows = []
    for i in range(fr['n']):
        rows.append([date_strs_h[i],
            round(rep_r['PE'][i], 2),
            round(rep_r['WU'][i] + rep_r['WL'][i] + rep_r['WD'][i], 1),
            round(rep_r['a'][i], 2),
            round(rep_r['R'][i], 3), round(rep_r['FR_series'][i], 3)])
    write_sheet(ws, hdr, rows, [14,10,10,10,10,10])

    # 三水源
    ws = wb_flood.create_sheet(f"三水源_{tag}"[:31])
    hdr = ['时间', 'R(mm)', 'S(mm)', 'FR', 'RS(mm)', 'RI(mm)', 'RG(mm)']
    rows = []
    for i in range(fr['n']):
        rows.append([date_strs_h[i],
            round(rep_r['R'][i], 3), round(rep_r['S'][i], 4),
            round(rep_r['FR_series'][i], 3),
            round(rep_r['RS'][i], 3), round(rep_r['RI'][i], 3), round(rep_r['RG'][i], 3)])
    write_sheet(ws, hdr, rows, [14,10,10,10,10,10,10])

    # 坡地汇流
    ws = wb_flood.create_sheet(f"坡地汇流_{tag}"[:31])
    hdr = ['时间', 'QS(m3/s)', 'QI(m3/s)', 'QG(m3/s)', 'QT(m3/s)']
    rows = []
    for i in range(fr['n']):
        rows.append([date_strs_h[i],
            round(rep_r['QS'][i], 3), round(rep_r['QI_series'][i], 3),
            round(rep_r['QG_series'][i], 3), round(rep_r['QT'][i], 3)])
    write_sheet(ws, hdr, rows, [14,12,12,12,12])

    # 河网汇流
    ws = wb_flood.create_sheet(f"河网汇流_{tag}"[:31])
    hdr = ['时间', 'QT_lag(m3/s)', 'Q_sim_before_routing(m3/s)']
    rows = []
    for i in range(fr['n']):
        rows.append([date_strs_h[i],
            round(rep_r['QT_lagged'][i], 3), round(rep_r['Q'][i], 3)])
    write_sheet(ws, hdr, rows, [14,14,18])

    # 河道汇流(马斯京根)
    if rep_nseg > 0:
        ws = wb_flood.create_sheet(f"马斯京根_{tag}"[:31])
        Q_in = rep_r['Q']
        Q_out = rep_sub['Q_routed']
        hdr = ['时间', 'Q_in(m3/s)', 'Q_out(m3/s)', '衰减(m3/s)']
        rows = []
        for i in range(fr['n']):
            rows.append([date_strs_h[i],
                round(Q_in[i], 3), round(Q_out[i], 3),
                round(Q_in[i] - Q_out[i], 3)])
        write_sheet(ws, hdr, rows, [14,14,14,14])

    # 最终出流
    ws = wb_flood.create_sheet(f"出流_{tag}"[:31])
    hdr = ['时间', 'P_areal(mm)', 'Q_obs(m3/s)', 'Q_sim(m3/s)', 'error(m3/s)']
    rows = []
    for i in range(fr['n']):
        err = fr['Q_sim'][i] - fr['Q_obs'][i]
        rows.append([date_strs_h[i],
            round(P_areal_h[i], 1),
            round(fr['Q_obs'][i], 2), round(fr['Q_sim'][i], 2), round(err, 3)])
    write_sheet(ws, hdr, rows, [14,10,12,12,12])

wb_flood.save(os.path.join(OUT_DIR, '次洪模型详细结果_v2.xlsx'))
print(f"  次洪模型Excel已保存: {os.path.join(OUT_DIR, '次洪模型详细结果.xlsx')}")

# ========== 日模型流量过程线 ==========
print("\n生成图表...")

# ========== 日模型：每年单独一张图 (双面板: 上雨量 + 下流量) ==========
DAILY_CHART_DIR = os.path.join(OUT_DIR, 'charts', 'daily')
os.makedirs(DAILY_CHART_DIR, exist_ok=True)

for idx, yr in enumerate(years_all):
    m = np.array([d.year == yr for d in dates_daily])
    d_idx = np.arange(m.sum())
    qo = Q_daily_obs[m]; qs = Q_daily_sim[m]
    p = P_areal[m]
    dy = daily_yearly[idx]
    tag = 'Calibration' if yr <= 1994 else 'Validation'

    fig, (ax_rain, ax_flow) = plt.subplots(2, 1, figsize=(16, 8),
        sharex=True, gridspec_kw={'height_ratios': [1, 2.5]})

    # 上图: 降雨柱状图
    ax_rain.bar(d_idx, p, color='#4A90D9', width=1.0, edgecolor='none', alpha=0.9)
    ax_rain.set_ylabel('Rainfall\n(mm)', fontsize=10, fontweight='bold')
    ax_rain.set_ylim(max(p.max() * 1.4, 10), 0)  # 倒置,顶部是0
    ax_rain.tick_params(labelsize=9)
    ax_rain.grid(axis='y', alpha=0.3, linestyle='--')
    ax_rain.set_title(f'Chengcun Daily Model — {yr} ({tag})', fontsize=13, fontweight='bold',
                      fontfamily='sans-serif')

    # 下图: 流量过程线
    ax_flow.plot(d_idx, qo, 'k-', linewidth=1.3, label='Observed', alpha=0.85)
    ax_flow.plot(d_idx, qs, 'r-', linewidth=1.0, label='Simulated', alpha=0.85)
    ax_flow.set_ylabel('Discharge\n(m$^3$/s)', fontsize=10, fontweight='bold')
    ax_flow.set_xlabel('Day', fontsize=10, fontweight='bold')
    ax_flow.tick_params(labelsize=9)
    ax_flow.grid(True, alpha=0.3, linestyle='--')
    ax_flow.legend(loc='upper right', fontsize=10, framealpha=0.9,
                   edgecolor='gray', fancybox=True)

    # 精度标注
    textstr = (f'BOw = {dy[2]:.4f}    '
               f'R$_{{err}}$ = {dy[3]:.1f}%    '
               f'R$_{{sim}}$ = {dy[4]:.0f} mm    '
               f'R$_{{obs}}$ = {dy[5]:.0f} mm')
    ax_flow.text(0.5, -0.18, textstr, transform=ax_flow.transAxes, fontsize=9,
                 ha='center', va='top',
                 bbox=dict(boxstyle='round,pad=0.4', facecolor='lightyellow',
                           edgecolor='gray', alpha=0.85))

    fig.tight_layout(rect=[0, 0.04, 1, 0.97])
    fname = f'daily_{yr}_{"rate" if yr<=1994 else "test"}.png'
    fig.savefig(os.path.join(DAILY_CHART_DIR, fname), dpi=180, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)

print(f"  日模型图已保存: {DAILY_CHART_DIR}/ ({len(years_all)} 张)")

# ========== 次洪模型：每场单独 (双面板: 上雨量 + 下流量) ==========
FLOOD_CHART_DIR = os.path.join(OUT_DIR, 'charts', 'flood')
os.makedirs(FLOOD_CHART_DIR, exist_ok=True)

for idx, fr in enumerate(flood_results):
    d_idx = np.arange(fr['n'])
    qo = fr['Q_obs']; qs = fr['Q_sim']
    fd = fr['fd']
    p = np.sum([fd['P_stations'][:, i] * sub_areas[i] for i in range(10)], axis=0) / F

    tag = "Calibration" if fr['is_rate'] else "Validation"
    status = "PASS" if fr['ok'] else "FAIL"

    fig, (ax_rain, ax_flow) = plt.subplots(2, 1, figsize=(16, 8),
        sharex=True, gridspec_kw={'height_ratios': [1, 2.5]})

    # 上图: 降雨柱状图
    ax_rain.bar(d_idx, p, color='#4A90D9', width=1.0, edgecolor='none', alpha=0.9)
    ax_rain.set_ylabel('Rainfall\n(mm)', fontsize=10, fontweight='bold')
    ax_rain.set_ylim(max(p.max() * 1.4, 10), 0)
    ax_rain.tick_params(labelsize=9)
    ax_rain.grid(axis='y', alpha=0.3, linestyle='--')
    ax_rain.set_title(f'Chengcun Flood #{fr["no"]} {tag}  —  {fr["date"].strftime("%Y-%m-%d")}  [{status}]',
                      fontsize=13, fontweight='bold', fontfamily='sans-serif')

    # 下图: 流量过程线
    ax_flow.plot(d_idx, qo, 'k-', linewidth=1.3, label='Observed', alpha=0.85)
    ax_flow.plot(d_idx, qs, 'r-', linewidth=1.0, label='Simulated', alpha=0.85)
    ax_flow.set_ylabel('Discharge\n(m$^3$/s)', fontsize=10, fontweight='bold')
    ax_flow.set_xlabel('Time (h)', fontsize=10, fontweight='bold')
    ax_flow.tick_params(labelsize=9)
    ax_flow.grid(True, alpha=0.3, linestyle='--')
    ax_flow.legend(loc='upper right', fontsize=10, framealpha=0.9,
                   edgecolor='gray', fancybox=True)

    # 精度标注
    textstr = (f'BOw = {fr["BOw"]:.4f}    '
               f'R$_{{err}}$ = {fr["R_err"]:.1f}%    '
               f'Qp$_{{err}}$ = {fr["Qp_err"]:.1f}%    '
               f'Tp$_{{err}}$ = {fr["Tp_err"]:.0f}h    '
               f'DC = {fr["DC"]:.3f}')
    ax_flow.text(0.5, -0.18, textstr, transform=ax_flow.transAxes, fontsize=9,
                 ha='center', va='top',
                 bbox=dict(boxstyle='round,pad=0.4',
                           facecolor='lightgreen' if fr['ok'] else 'mistyrose',
                           edgecolor='gray', alpha=0.85))

    fig.tight_layout(rect=[0, 0.04, 1, 0.97])
    fname = f'flood_{fr["no"]:02d}_{fr["date"].strftime("%Y%m%d")}.png'
    fig.savefig(os.path.join(FLOOD_CHART_DIR, fname), dpi=180, bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)

print(f"  次洪模型图已保存: {FLOOD_CHART_DIR}/ ({len(flood_results)} 张)")

# ========== 汇总 ==========
print("\n" + "=" * 60)
print("全部输出完成!")

# 率定/检验分开统计
rate_bo = [f['BOw'] for f in flood_results if f['is_rate']]
test_bo = [f['BOw'] for f in flood_results if not f['is_rate']]
rate_r = [f['R_err'] for f in flood_results if f['is_rate']]
test_r = [f['R_err'] for f in flood_results if not f['is_rate']]
rate_q = [f['Qp_err'] for f in flood_results if f['is_rate']]
test_q = [f['Qp_err'] for f in flood_results if not f['is_rate']]
rate_ok = sum(1 for f in flood_results if f['is_rate'] and f['ok'])
test_ok = sum(1 for f in flood_results if not f['is_rate'] and f['ok'])

rate_d_bo = [d[2] for d in daily_yearly if d[0] <= 1994]
test_d_bo = [d[2] for d in daily_yearly if d[0] >= 1995]
rate_d_re = [d[3] for d in daily_yearly if d[0] <= 1994]
test_d_re = [d[3] for d in daily_yearly if d[0] >= 1995]

print(f"\nDaily model:")
print(f"  Rate(1989-1994): BOw={np.mean(rate_d_bo):.4f}  R_err={np.mean(rate_d_re):.1f}%")
print(f"  Test(1995-1996): BOw={np.mean(test_d_bo):.4f}  R_err={np.mean(test_d_re):.1f}%")
print(f"  Year OK: {sum(1 for d in daily_yearly if d[3]<=20)}/{len(daily_yearly)}")

print(f"\n次洪模型:")
print(f"  率定: BOw={np.mean(rate_bo):.4f}  R_err={np.mean(rate_r):.1f}%  Qp_err={np.mean(rate_q):.1f}%  达标{rate_ok}/{len(rate_r)}")
print(f"  检验: BOw={np.mean(test_bo):.4f}  R_err={np.mean(test_r):.1f}%  Qp_err={np.mean(test_q):.1f}%  达标{test_ok}/{len(test_r)}")
print(f"  总达标: {rate_ok+test_ok}/{len(flood_results)}")

print(f"\n输出目录: {OUT_DIR}")
print("  /日模型详细结果.xlsx")
print("  /次洪模型详细结果.xlsx")
print("  /charts/daily/ (8张逐年流量过程线)")
print("  /charts/flood/ (15张场次洪水流量过程线)")
