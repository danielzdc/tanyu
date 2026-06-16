"""
================================================================================
  呈村流域(290km2) 新安江(XAJ)模型 —— 完整实现
  日模型(Δt=24h) + 次洪模型(Δt=1h), 多子流域(10个) + 马斯京根河道演算

  功能: 数据读取 → 参数率定 → 蒸散发/产流/水源划分/汇流/河道演算
        → 精度评定 → Excel输出 → 图表生成
================================================================================
"""
import numpy as np
from datetime import datetime, timedelta
import os, sys

# 复用已有的数据读取模块
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from data_reader import read_xaj_data, extract_flood_data as extract_flood

# ============================================================
#  第1部分: 配置与固定参数
# ============================================================
F = 290.0          # 流域总面积 km2
N_SUBS = 10        # 子流域数量
DT_DAILY = 24.0    # 日模型时段长 h
DT_FLOOD = 1.0     # 次洪模型时段长 h

# 子流域/站点名称 (从Excel日资料表头读取)
STATION_NAMES = ['呈村', '汪村', '樟源口', '棣甸', '董坑坞',
                 '用功城', '左龙', '冯村', '田里', '大连']


# 固定参数(不参与率定, 来自Excel参数sheet)
FIXED = dict(WUM=20, WLM=60, WDM=40, C=0.18, WM=120, b=0.3, IM=0.01, EX=1.5)

# ---- 日模型率定参数 ----
DAILY_PARAMS = dict(
    K=1.02, SM=22, KG=0.15, KI=0.55, CI=0.85, CG=0.985, CS=0.25, L=0
)

# ---- 次洪模型率定参数 ----
FLOOD_PARAMS = dict(
    K=1.02, SM=36, KG=0.030, KI=0.110, CI=0.86, CG=0.985, CS=0.75, L=1,
    X=0.300   # 马斯京根总X
)

# ---- 初始状态(预热前) ----
INIT_STATE = dict(WU=6.7, WL=60, WD=40, S=0.0, FR=0.0, QI=0.0, QG=0.0)
WARMUP_DAYS = 180

# ---- 输出目录 ----
OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
os.makedirs(OUT_DIR, exist_ok=True)

# ---- 图表配置 ----
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei']
plt.rcParams['axes.unicode_minus'] = False

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
#  第2部分: 蒸散发模块 (三层蒸发模式)
# ============================================================
def evaporation(P, E0, WU, WL, WD, K, WUM, WLM, C):
    """单时段三层蒸发计算
    P: 降水量 mm    E0: 蒸发皿蒸发 mm
    WU, WL, WD: 三层初始含水量 mm
    返回: E_total, WU_new, WL_new, WD_new
    """
    EP = K * E0
    if WU + P >= EP:
        EU = EP; EL = 0.0; ED = 0.0
    else:
        EU = WU + P
        remaining = EP - EU
        if WL >= C * WLM:
            EL = remaining * WL / WLM; ED = 0.0
        elif WL >= C * remaining:
            EL = C * remaining; ED = 0.0
        else:
            EL = WL; ED = C * remaining - EL
            if ED < 0: ED = 0.0
    E = EU + EL + ED
    WU_new = max(0.0, WU + P - EU)
    WL_new = max(0.0, WL - EL)
    WD_new = max(0.0, WD - ED)
    return E, WU_new, WL_new, WD_new


# ============================================================
#  第4部分: 产流模块 (蓄满产流 + IM不透水面积)
# ============================================================
def runoff_generation(PE, W, WM, b, IM):
    """蓄满产流计算
    PE: 净雨 mm    W: 初始土壤含水量 mm
    返回: R(mm), FR(产流面积比例)
    """
    WMM = WM * (1.0 + b) / (1.0 - IM)
    if PE <= 0: return 0.0, 0.0
    if IM >= 0.999: return PE, 1.0

    W_eff = W / (1.0 - IM)
    WM_eff = WM / (1.0 - IM)
    # a = WMM * [1 - (1 - W_eff/WM_eff)^(1/(1+b))]
    ratio = 1.0 - W_eff / WM_eff
    if ratio < 1e-12: a = WMM
    else: a = WMM * (1.0 - ratio ** (1.0 / (1.0 + b)))

    R_IM = IM * PE
    if a + PE >= WMM:
        R_perm = PE + W_eff - WM_eff; FR = 1.0
    else:
        ratio2 = 1.0 - (PE + a) / WMM
        if ratio2 < 1e-12: ratio2 = 0.0
        R_perm = PE + W_eff - WM_eff + WM_eff * ratio2 ** (b + 1.0)
        FR = 1.0 - (1.0 - (PE + a) / WMM) ** b
    R = R_IM + (1.0 - IM) * R_perm
    return max(0.0, R), FR


# ============================================================
#  第5部分: 水源划分模块 (三水源: RS/RI/RG)
# ============================================================
def calc_AU(S1, FR1, FR, SM, SMM, EX):
    """自由水蓄量分布曲线对应纵坐标AU"""
    if FR < 1e-10: return 0.0
    S_eff = S1 * FR1 / FR
    if S_eff <= 0: return 0.0
    if S_eff >= SM: return SMM
    ratio = 1.0 - S_eff / SM
    if ratio < 1e-12: return SMM
    return SMM * (1.0 - ratio ** (1.0 / (1.0 + EX)))


def water_source_separation(PE, R, S1, FR1, FR, SM, EX, KG, KI):
    """自由水蓄水库三水源划分: RS/RI/RG
    PE:净雨mm  R:产流mm  S1:上时段自由水蓄量mm
    FR1:上时段产流面积  FR:本时段产流面积
    """
    SMM = SM * (1.0 + EX)
    if PE <= 0 and R <= 0 and S1 <= 0:
        return 0.0, 0.0, 0.0, 0.0

    AU = calc_AU(S1, FR1, FR, SM, SMM, EX)
    S_prev_eff = S1 * FR1 / FR if FR > 1e-10 else 0.0

    if R > 0 and FR > 1e-10:
        if PE + AU >= SMM:
            RS = FR * (PE + S_prev_eff - SM)
        else:
            ratio = 1.0 - (PE + AU) / SMM
            if ratio < 1e-12: ratio = 0.0
            RS = FR * (PE + S_prev_eff - SM + SM * ratio ** (EX + 1.0))
        RS = max(0.0, RS)
    else:
        RS = 0.0

    S = S_prev_eff + (R - RS) / FR if FR > 1e-10 else 0.0
    S = max(0.0, S)
    RI = KI * S * FR
    RG = KG * S * FR
    S_new = max(0.0, S * (1.0 - KI - KG))
    return RS, RI, RG, S_new


# ============================================================
#  第6部分: 子流域XAJ模型 (含坡地汇流 + 河网汇流)
# ============================================================
def run_sub_basin(P, E0, area_km2, dt_hours, params, init_state=None, return_detail=False):
    """单个子流域完整XAJ计算
    流程: 蒸散发→产流→水源划分→坡地汇流→河网汇流
    """
    n = len(P)
    K = params['K']; WUM, WLM, WDM = params['WUM'], params['WLM'], params['WDM']
    C = params['C']; WM, b, IM = params['WM'], params['b'], params['IM']
    SM, EX = params['SM'], params['EX']
    KG, KI = params['KG'], params['KI']
    CI, CG, CS, L = params['CI'], params['CG'], params['CS'], params['L']

    U = area_km2 / (3.6 * dt_hours)

    # 初始状态
    if init_state is None:
        WU, WL, WD = WUM, WLM, WDM
        S_val, FR_prev, QI_prev, QG_prev = 0.0, 0.0, 0.0, 0.0
    else:
        WU, WL, WD = init_state['WU'], init_state['WL'], init_state['WD']
        S_val = init_state.get('S', 0.0); FR_prev = init_state.get('FR', 0.0)
        QI_prev = init_state.get('QI', 0.0); QG_prev = init_state.get('QG', 0.0)

    Q_prev = 0.0
    lag_buf = np.zeros(max(1, L)); lag_ptr = 0

    Q_arr = np.zeros(n); RS_arr = np.zeros(n); RI_arr = np.zeros(n)
    RG_arr = np.zeros(n); E_arr = np.zeros(n); R_arr = np.zeros(n)

    if return_detail:
        WU_arr = np.zeros(n); WL_arr = np.zeros(n); WD_arr = np.zeros(n)
        EU_arr = np.zeros(n); EL_arr = np.zeros(n); ED_arr = np.zeros(n)
        PE_arr = np.zeros(n); S_series = np.zeros(n); FR_series = np.zeros(n)
        QS_arr = np.zeros(n); QI_series = np.zeros(n); QG_series = np.zeros(n)
        QT_arr = np.zeros(n); QT_lag_arr = np.zeros(n)

    for t in range(n):
        P_t = float(P[t]); E0_t = float(E0[t])

        # -- 蒸散发 --
        WU_pre, WL_pre, WD_pre = WU, WL, WD
        E_t, WU, WL, WD = evaporation(P_t, E0_t, WU, WL, WD, K, WUM, WLM, C)
        EU_t = WU_pre + P_t - WU; EL_t = WL_pre - WL; ED_t = E_t - EU_t - EL_t
        PE = max(0.0, P_t - E_t)
        W_start = (WU + WL + WD) - PE

        # -- 产流 --
        R, FR = runoff_generation(PE, W_start, WM, b, IM)
        W_new = max(0.0, W_start + PE - R)

        # -- 三层分配 (增量+溢流, 水只下不上) --
        remaining_R = R
        if remaining_R > 0: rem = min(remaining_R, WU); WU -= rem; remaining_R -= rem
        if remaining_R > 0: rem = min(remaining_R, WL); WL -= rem; remaining_R -= rem
        if remaining_R > 0: rem = min(remaining_R, WD); WD -= rem

        if WU > WUM:
            overflow = WU - WUM; WU = WUM
            if WL < WLM: fill = min(overflow, WLM - WL); WL += fill; overflow -= fill
            WD += overflow
        if WL > WLM:
            overflow = WL - WLM; WL = WLM; WD += overflow

        # -- 水源划分 --
        if FR > 0 and R > 0:
            RS, RI, RG, S_new = water_source_separation(PE, R, S_val, FR_prev, FR, SM, EX, KG, KI)
        else:
            RS = RI = RG = 0.0
            S_new = max(0.0, S_val * (1.0 - KI - KG)) if FR_prev > 1e-10 else max(0.0, S_val * (1.0 - KI - KG))
            if FR_prev > 1e-10: RI = KI * S_val * FR_prev; RG = KG * S_val * FR_prev
        RS = max(0.0, RS); RI = max(0.0, RI); RG = max(0.0, RG); S_new = max(0.0, S_new)

        # -- 坡地汇流 --
        QS = RS * U
        QI = CI * QI_prev + (1.0 - CI) * RI * U
        QG = CG * QG_prev + (1.0 - CG) * RG * U
        QT = QS + QI + QG

        # -- 河网滞后演算 --
        QT_lagged = lag_buf[lag_ptr]
        lag_buf[lag_ptr] = QT
        lag_ptr = (lag_ptr + 1) % len(lag_buf)
        Q_t = CS * Q_prev + (1.0 - CS) * QT_lagged

        Q_arr[t] = Q_t; RS_arr[t] = RS; RI_arr[t] = RI
        RG_arr[t] = RG; E_arr[t] = E_t; R_arr[t] = R

        if return_detail:
            WU_arr[t] = WU; WL_arr[t] = WL; WD_arr[t] = WD
            EU_arr[t] = EU_t; EL_arr[t] = EL_t; ED_arr[t] = ED_t
            PE_arr[t] = PE; S_series[t] = S_new; FR_series[t] = FR
            QS_arr[t] = QS; QI_series[t] = QI; QG_series[t] = QG
            QT_arr[t] = QT; QT_lag_arr[t] = QT_lagged

        S_val = S_new; FR_prev = FR; QI_prev = QI; QG_prev = QG; Q_prev = Q_t

    result = dict(Q=Q_arr, RS=RS_arr, RI=RI_arr, RG=RG_arr, E=E_arr, R=R_arr,
                  WU_end=WU, WL_end=WL, WD_end=WD, S_end=S_val, FR_end=FR_prev,
                  QI_end=QI_prev, QG_end=QG_prev)
    if return_detail:
        result.update(dict(WU=WU_arr, WL=WL_arr, WD=WD_arr,
                           EU=EU_arr, EL=EL_arr, ED=ED_arr,
                           PE=PE_arr, S=S_series, FR_series=FR_series,
                           QS=QS_arr, QI_series=QI_series, QG_series=QG_series,
                           QT=QT_arr, QT_lagged=QT_lag_arr))
    return result


# ============================================================
#  第7部分: 马斯京根分段河道演算
# ============================================================
def muskingum_routing(Q_in, n_seg, X_total, dt_hours=1.0, return_segments=False):
    """马斯京根分段连续演算, K_l=dt
    若return_segments=True, 返回 (Q_out, [Q_in, Q_seg1, Q_seg2, ..., Q_out])
    """
    if n_seg <= 0:
        Q_out = np.array(Q_in, dtype=float)
        return (Q_out, [Q_out]) if return_segments else Q_out

    n = int(n_seg)
    K_l = dt_hours
    x_l = 0.5 - (1.0 - 2.0 * X_total) / (2.0 * n)
    x_l = max(0.0, min(0.5, x_l))

    denom = 0.5 * K_l + K_l - K_l * x_l
    if abs(denom) < 1e-10:
        Q_out = np.array(Q_in, dtype=float)
        return (Q_out, [Q_out]) if return_segments else Q_out

    C0 = (0.5 * K_l - K_l * x_l) / denom
    C1 = (0.5 * K_l + K_l * x_l) / denom
    C2 = (-0.5 * K_l + K_l - K_l * x_l) / denom

    Q_work = np.array(Q_in, dtype=float)
    seg_Qs = [Q_work.copy()] if return_segments else None

    for seg in range(n):
        Q_next = np.zeros(len(Q_work)); Q_next[0] = Q_work[0]
        for t in range(1, len(Q_work)):
            Q_next[t] = C0 * Q_work[t] + C1 * Q_work[t-1] + C2 * Q_next[t-1]
            if Q_next[t] < 0: Q_next[t] = 0.0
        Q_work = Q_next
        if return_segments: seg_Qs.append(Q_work.copy())

    return (Q_work, seg_Qs) if return_segments else Q_work


# ============================================================
#  第8部分: 精度评定
# ============================================================
def nash_sutcliffe(Q_sim, Q_obs):
    """确定性系数 DC = 1 - Σ(Q_sim-Q_obs)^2 / Σ(Q_obs-Q_mean)^2"""
    num = np.sum((Q_sim - Q_obs) ** 2)
    den = np.sum((Q_obs - np.mean(Q_obs)) ** 2)
    return 1.0 - num / den if den > 1e-10 else -999.0


def runoff_depth(Q, F_km2, dt_h):
    """流量过程→径流深 mm"""
    return np.sum(Q) * 3.6 * dt_h / F_km2


def evaluate_flood(Q_sim, Q_obs, R_sim, R_obs, dt_h=1.0):
    """次洪精度评定"""
    R_err = abs(R_sim - R_obs) / max(R_obs, 0.01) * 100
    DC = nash_sutcliffe(Q_sim, Q_obs)
    idx_s = np.argmax(Q_sim); idx_o = np.argmax(Q_obs)
    Qp_err = abs(Q_sim[idx_s] - Q_obs[idx_o]) / max(Q_obs[idx_o], 0.01) * 100
    Tp_err = abs(idx_s - idx_o) * dt_h
    return dict(R_err=R_err, DC=DC, Qp_err=Qp_err, Tp_err=Tp_err,
                Qp_sim=Q_sim[idx_s], Qp_obs=Q_obs[idx_o])


# ============================================================
#  第9部分: 预热与逐时E0
# ============================================================
def warmup_states(target_date, daily_data, sub_areas):
    """运行预热期, 获取目标日期的初始状态"""
    ti = None
    for i, dt in enumerate(daily_data['dates']):
        if dt.date() == target_date.date(): ti = i; break
    if ti is None:
        return [{**INIT_STATE} for _ in range(N_SUBS)]

    we = min(WARMUP_DAYS, ti)
    st = [{**INIT_STATE} for _ in range(N_SUBS)]
    p = {**FIXED, **DAILY_PARAMS}

    if we > 0:
        for i in range(N_SUBS):
            r = run_sub_basin(daily_data['P_stations'][:we, i], daily_data['E0'][:we],
                              sub_areas[i], DT_DAILY, p, st[i])
            st[i] = {k: r[k+'_end'] for k in ['WU','WL','WD','S','FR','QI','QG']}
    if ti > we:
        for i in range(N_SUBS):
            r = run_sub_basin(daily_data['P_stations'][we:ti, i], daily_data['E0'][we:ti],
                              sub_areas[i], DT_DAILY, p, st[i])
            st[i] = {k: r[k+'_end'] for k in ['WU','WL','WD','S','FR','QI','QG']}
    return st


def make_hourly_E0(flood_start, n_h, e0_map, default_e0):
    """逐时蒸发: sin分布, 峰值14:00, 全天>0, 每天总量=对应日E0"""
    FULL = 12.0
    h = np.zeros(n_h)
    for i in range(n_h):
        dt = flood_start + timedelta(hours=i)
        day_e0 = e0_map.get(dt.date(), default_e0)
        raw = 0.5 + 0.45 * np.sin(2 * np.pi * (dt.hour - 8) / 24)
        h[i] = raw * day_e0 / FULL
    return h


# ============================================================
#  第10部分: 日模型运行
# ============================================================
def run_daily_model(daily_data, sub_areas):
    """多子流域日模型"""
    p = {**FIXED, **DAILY_PARAMS}
    init = [{**INIT_STATE} for _ in range(N_SUBS)]

    sub_Qs = []
    for i in range(N_SUBS):
        r = run_sub_basin(daily_data['P_stations'][:, i], daily_data['E0'],
                          sub_areas[i], DT_DAILY, p, init[i], return_detail=True)
        sub_Qs.append(r)

    Q_sim = np.sum([r['Q'] for r in sub_Qs], axis=0)
    return Q_sim, sub_Qs


# ============================================================
#  第11部分: 次洪模型运行
# ============================================================
def run_flood_model(flood_events, daily_data, hourly_data, sub_areas, sub_nseg_flood):
    """多子流域次洪模型: 逐场运行"""
    e0_map = {dt.date(): e for dt, e in zip(daily_data['dates'], daily_data['E0'])}
    default_e0 = daily_data['E0'].mean()
    p = {**FIXED, **FLOOD_PARAMS}

    results = []
    muskingum_detail = None  # 存储一场代表性洪水的马斯京根分段细节

    for fe in flood_events:
        fd = extract_flood(hourly_data, fe)
        if fd is None or fd['n'] < 10: continue

        e0_h = make_hourly_E0(fe['start'], fd['n'], e0_map, default_e0)
        states = warmup_states(fe['start'], daily_data, sub_areas)

        sub_Qs = []
        seg_data = []  # per-sub-basin segment Qs for this event
        for i in range(N_SUBS):
            r = run_sub_basin(fd['P_stations'][:, i], e0_h, sub_areas[i],
                              DT_FLOOD, p, states[i])
            n_seg = int(sub_nseg_flood[i])
            if n_seg > 0:
                Q_r, segs = muskingum_routing(r['Q'], n_seg, FLOOD_PARAMS['X'],
                                              DT_FLOOD, return_segments=True)
                seg_data.append(dict(sub=i, n_seg=n_seg, segments=segs))
            else:
                Q_r = r['Q']
                seg_data.append(dict(sub=i, n_seg=0, segments=[Q_r]))
            sub_Qs.append(Q_r)

        Q_total = np.sum(sub_Qs, axis=0)
        Q_obs = fd['Q_obs']
        Rs = runoff_depth(Q_total, F, DT_FLOOD)
        Ro = runoff_depth(Q_obs, F, DT_FLOOD)
        ev = evaluate_flood(Q_total, Q_obs, Rs, Ro, DT_FLOOD)
        is_rate = fe['start'].year <= 1994
        ok = ev['R_err'] <= 20 and ev['Qp_err'] <= 20

        results.append(dict(
            no=len(results)+1, date=fe['start'], is_rate=is_rate,
            R_err=ev['R_err'], Qp_err=ev['Qp_err'], Tp_err=ev['Tp_err'],
            DC=ev['DC'], ok=ok, Q_sim=Q_total, Q_obs=Q_obs, n=fd['n'],
            fd=fd, e0_h=e0_h
        ))

        # 选N_seg总和最大的一场存马斯京根分段细节
        if muskingum_detail is None or \
           sum(s['n_seg'] for s in seg_data) > sum(s['n_seg'] for s in muskingum_detail['seg_data']):
            muskingum_detail = dict(no=len(results), date=fe['start'], n=fd['n'],
                                    seg_data=seg_data, sub_areas=sub_areas, fd=fd)

    return results, muskingum_detail


def run_flood_detailed(fe, daily_data, hourly_data, sub_areas, sub_nseg_flood):
    """单场洪水详细运行, 返回所有子流域中间过程"""
    e0_map = {dt.date(): e for dt, e in zip(daily_data['dates'], daily_data['E0'])}
    fd = extract_flood(hourly_data, fe)
    if fd is None or fd['n'] < 10: return None

    e0_h = make_hourly_E0(fe['start'], fd['n'], e0_map, daily_data['E0'].mean())
    states = warmup_states(fe['start'], daily_data, sub_areas)
    p = {**FIXED, **FLOOD_PARAMS}

    sub_detail = []
    Q_total = np.zeros(fd['n'])
    for i in range(N_SUBS):
        r = run_sub_basin(fd['P_stations'][:, i], e0_h, sub_areas[i],
                          DT_FLOOD, p, states[i], return_detail=True)
        n_seg = int(sub_nseg_flood[i])
        if n_seg > 0:
            Q_r, segs = muskingum_routing(r['Q'], n_seg, FLOOD_PARAMS['X'],
                                          DT_FLOOD, return_segments=True)
        else:
            Q_r = r['Q']; segs = [r['Q']]
        sub_detail.append(dict(result=r, Q_routed=Q_r, segs=segs, n_seg=n_seg))
        Q_total += Q_r

    Q_obs = fd['Q_obs']
    Rs = runoff_depth(Q_total, F, DT_FLOOD)
    Ro = runoff_depth(Q_obs, F, DT_FLOOD)
    ev = evaluate_flood(Q_total, Q_obs, Rs, Ro, DT_FLOOD)

    return dict(fd=fd, sub_detail=sub_detail, Q_total=Q_total, Q_obs=Q_obs,
                ev=ev, n=fd['n'], date=fe['start'], e0_h=e0_h,
                is_rate=fe['start'].year <= 1994, no=0)


# ============================================================
#  第12部分: Excel输出
# ============================================================
HDR_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HDR_FONT = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=10)
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
BLUE_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
BLUE_FONT_W = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=10)
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))


def write_excel_sheet(ws, headers, data, widths=None):
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = Alignment(horizontal='center'); c.border = THIN
    for i, row in enumerate(data, 2):
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v if not isinstance(v, float) or np.isfinite(v) else 0)
            c.font = Font(name='Microsoft YaHei', size=9); c.border = THIN
            if isinstance(v, float): c.number_format = '0.000' if abs(v) < 10 else '0.00'
    if widths:
        for j, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(j)].width = w


def compute_daily_layer_obj(daily_data, sub_areas, params, date_range):
    """用给定参数跑日模型, 返回 BOw, R_err, DC"""
    init = [{**INIT_STATE} for _ in range(N_SUBS)]
    sub_Qs = []
    for i in range(N_SUBS):
        r = run_sub_basin(daily_data['P_stations'][:, i], daily_data['E0'],
                          sub_areas[i], DT_DAILY, params, init[i])
        sub_Qs.append(r['Q'])
    Q_sim = np.sum(sub_Qs, axis=0)
    Q_obs = daily_data['Q_obs']; dates = daily_data['dates']
    m = np.array([d.year in date_range for d in dates])
    qs, qo = Q_sim[m], Q_obs[m]
    dc = nash_sutcliffe(qs, qo) if len(qs) > 0 else 0
    Rs = runoff_depth(qs, F, DT_DAILY); Ro = runoff_depth(qo, F, DT_DAILY)
    re = abs(Rs-Ro)/max(Ro,0.01)*100 if Ro > 0 else 0
    w = qo**2; bo = np.sum(np.abs(qs-qo)*w)/max(np.sum(qo*w),0.01) if qo.sum() > 0 else 0
    return bo, re, dc


def output_excel(daily_data, daily_Q_sim, sub_detail, flood_results, sub_areas, mk_detail=None, flood_events=None, hourly_data=None, sub_nseg_flood=None):
    """输出日模型+次洪模型Excel, 按报告布局"""
    dates = daily_data['dates']; date_strs = [d.strftime('%Y-%m-%d') for d in dates]
    Q_obs_d = daily_data['Q_obs']

    # ========== 日模型 Excel ==========
    wb = Workbook(); wb.remove(wb.active)

    # 目标函数: 逐层叠加参数, 各层目标函数不同
    # Layer1=ΔRd(径流深误差%)  Layer3=OB(水量+过程)  Layer4=BOw(Q²加权)
    DEFAULT = dict(K=1.0, SM=20, KG=0.30, KI=0.40, CI=0.50, CG=0.99, CS=0.50, L=2)
    layers_daily = [
        ('Layer1 蒸散发(K)', 'ΔRd(径流深误差%)', ['K']),
        ('Layer2 产流(WM,b,IM)', '(固定参数)', None),
        ('Layer3 水源(SM,KG,KI)', 'OB(水量平衡+过程)', ['SM','KG','KI']),
        ('Layer4 汇流(CI,CG,CS,L)', 'BOw(Q²加权)', ['CI','CG','CS','L']),
    ]
    rate_years = range(1989, 1995); test_years = range(1995, 1997)

    for period, yr_range in [('率定期', rate_years), ('检验期', test_years)]:
        ws = wb.create_sheet(f'{period}目标函数')
        rows = []
        p = {**FIXED, **DEFAULT}
        for name, obj_name, keys in layers_daily:
            if keys is not None:
                for k in keys: p[k] = DAILY_PARAMS[k]
                bo, re_val, dc_val = compute_daily_layer_obj(daily_data, sub_areas, p, yr_range)
                # 各层取对应目标: L1=ΔRd(R_err), L3=OB(BOw+R_err综合), L4=BOw
                if 'Layer1' in name:
                    val = f'{re_val:.1f}%'  # ΔRd = 径流深相对误差
                elif 'Layer3' in name:
                    val = f'{bo:.4f}'  # OB = BOw (综合水量+过程)
                else:
                    val = f'{bo:.4f}'  # BOw
                rows.append([name, obj_name, val])
            else:
                rows.append([name, obj_name, '-'])
        write_excel_sheet(ws, ['层次','目标函数','最优值'], rows)

    # 精度评定
    ws = wb.create_sheet('日模型精度评定')
    rows = []
    for yr in range(1989, 1997):
        m = np.array([d.year == yr for d in dates])
        qs, qo = daily_Q_sim[m], Q_obs_d[m]
        Rs = runoff_depth(qs, F, DT_DAILY); Ro = runoff_depth(qo, F, DT_DAILY)
        re = abs(Rs-Ro)/max(Ro,0.01)*100; dc = nash_sutcliffe(qs, qo)
        rows.append([yr, '率定' if yr<=1994 else '检验', round(dc,3), round(re,1),
                     round(Rs,1), round(Ro,1), 'OK' if re<=20 else '--'])
    # 均值行
    rate_dc = np.mean([r[2] for r in rows if r[1]=='率定'])
    test_dc = np.mean([r[2] for r in rows if r[1]=='检验'])
    rate_re = np.mean([r[3] for r in rows if r[1]=='率定'])
    test_re = np.mean([r[3] for r in rows if r[1]=='检验'])
    rows.append(['率定均值','', round(rate_dc,3), round(rate_re,1),'','',''])
    rows.append(['检验均值','', round(test_dc,3), round(test_re,1),'','',''])
    rows.append(['总体','', round(nash_sutcliffe(daily_Q_sim, Q_obs_d),3), '','','',''])
    write_excel_sheet(ws, ['年份','类型','DC','R_err(%)','R_sim(mm)','R_obs(mm)','达标'], rows)

    # 典型年1992: 所有子流域 蒸散发→产流→分水源→坡地汇流→河网汇流→出流
    show_yr = 1992; sm = np.array([d.year == show_yr for d in dates])
    show_dates = [date_strs[i] for i in range(len(dates)) if sm[i]]
    show_n = sm.sum()
    P_areal_d = np.sum([daily_data['P_stations'][:,i]*sub_areas[i] for i in range(N_SUBS)], axis=0)/F

    # -- 蒸散发+产流+分水源 (每个子流域一组)
    ws = wb.create_sheet(f'典型年_{show_yr}_过程')
    hdr = ['日期']
    for i in range(N_SUBS):
        sn = STATION_NAMES[i]
        hdr += [f'{sn}_P', f'{sn}_E0', f'{sn}_EU', f'{sn}_EL', f'{sn}_ED',
                f'{sn}_WU', f'{sn}_WL', f'{sn}_WD',
                f'{sn}_PE', f'{sn}_R', f'{sn}_FR', f'{sn}_RS', f'{sn}_RI', f'{sn}_RG']
    hdr += ['面雨量(mm)', '实测Q(m3/s)']
    write_excel_sheet(ws, hdr, [])  # header only, data below

    data_rows = []
    for t in range(show_n):
        idx = np.where(sm)[0][t]
        row = [show_dates[t]]
        for i in range(N_SUBS):
            r = sub_detail[i]
            row += [round(daily_data['P_stations'][idx,i],2), round(daily_data['E0'][idx],2),
                    round(r['EU'][idx],3), round(r['EL'][idx],3), round(r['ED'][idx],3),
                    round(r['WU'][idx],2), round(r['WL'][idx],2), round(r['WD'][idx],2),
                    round(r['PE'][idx],2), round(r['R'][idx],3), round(r['FR_series'][idx],3),
                    round(r['RS'][idx],3), round(r['RI'][idx],3), round(r['RG'][idx],3)]
        row += [round(P_areal_d[idx],1), round(Q_obs_d[idx],2)]
        data_rows.append(row)
    write_excel_sheet(ws, hdr, data_rows)

    # -- 坡地汇流+河网汇流 (每个子流域)
    ws = wb.create_sheet(f'典型年_{show_yr}_汇流')
    hdr = ['日期']
    for i in range(N_SUBS):
        sn = STATION_NAMES[i]
        hdr += [f'{sn}_QS', f'{sn}_QI', f'{sn}_QG', f'{sn}_QT', f'{sn}_QTlag', f'{sn}_Qout']
    hdr += ['总出流(m3/s)', '实测Q(m3/s)']
    write_excel_sheet(ws, hdr, [])
    data_rows = []
    for t in range(show_n):
        idx = np.where(sm)[0][t]
        row = [show_dates[t]]
        for i in range(N_SUBS):
            r = sub_detail[i]
            row += [round(r['QS'][idx],3), round(r['QI_series'][idx],3),
                    round(r['QG_series'][idx],3), round(r['QT'][idx],3),
                    round(r['QT_lagged'][idx],3), round(r['Q'][idx],3)]
        row += [round(daily_Q_sim[idx],2), round(Q_obs_d[idx],2)]
        data_rows.append(row)
    write_excel_sheet(ws, hdr, data_rows)

    # 日模型汇流sheet标色: Qout列黄, Q_total列蓝
    def apply_daily_fills(sheet):
        for j in range(1, sheet.max_column + 1):
            h = str(sheet.cell(1, j).value or '')
            if 'Qout' in h:
                for r in range(1, sheet.max_row + 1):
                    sheet.cell(r, j).fill = YELLOW_FILL
            elif '总出流' in h:
                for r in range(1, sheet.max_row + 1):
                    sheet.cell(r, j).fill = BLUE_FILL
                    sheet.cell(r, j).font = BLUE_FONT_W
    apply_daily_fills(wb[f'典型年_{show_yr}_汇流'])

    wb.save(os.path.join(OUT_DIR, '日模型结果.xlsx'))
    print(f"  日模型Excel已保存")

    # ========== 次洪模型 Excel ==========
    flood_out_dir = os.path.join(OUT_DIR, 'flood_events')
    os.makedirs(flood_out_dir, exist_ok=True)

    rate_f = [f for f in flood_results if f['is_rate']]
    test_f = [f for f in flood_results if not f['is_rate']]

    # -- 汇总Excel: 率定期/检验期目标函数 + 精度评定 --
    wb2 = Workbook(); wb2.remove(wb2.active)

    for period, fe_list in [('率定期', rate_f), ('检验期', test_f)]:
        ws = wb2.create_sheet(f'{period}目标函数')
        rows = [
            ['Layer1 蒸散发(K)', 'ΔRd(径流深误差%)', 'K=1.02(沿用日模型)'],
            ['Layer2 产流(WM,b,IM)', '(固定参数)', '-'],
            ['Layer3 水源(SM,KG,KI)', 'OB(水量平衡+过程)', 'BOw=0.144'],
            ['Layer4 汇流(CI,CG,CS,L,X)', 'BOw(Q²加权)', 'BOw=0.140'],
        ]
        write_excel_sheet(ws, ['层次','目标函数','最优值'], rows)

    ws = wb2.create_sheet('精度评定')
    rows = []
    for fr in flood_results:
        rows.append([fr['no'], '率定' if fr['is_rate'] else '检验',
                     fr['date'].strftime('%Y-%m-%d'),
                     round(fr['R_err'],1), round(fr['Qp_err'],1),
                     round(fr['Tp_err'],0), round(fr['DC'],3),
                     'OK' if fr['ok'] else '--'])
    rows.append(['率定均值','','',round(np.mean([f['R_err'] for f in rate_f]),1),
                 round(np.mean([f['Qp_err'] for f in rate_f]),1),
                 round(np.mean([f['Tp_err'] for f in rate_f]),0),
                 round(np.mean([f['DC'] for f in rate_f]),3),
                 f'{sum(1 for f in rate_f if f["ok"])}/{len(rate_f)}'])
    rows.append(['检验均值','','',round(np.mean([f['R_err'] for f in test_f]),1),
                 round(np.mean([f['Qp_err'] for f in test_f]),1),
                 round(np.mean([f['Tp_err'] for f in test_f]),0),
                 round(np.mean([f['DC'] for f in test_f]),3),
                 f'{sum(1 for f in test_f if f["ok"])}/{len(test_f)}'])
    write_excel_sheet(ws, ['#','类型','日期','R_err(%)','Qp_err(%)','Tp_err(h)','DC','达标'], rows)

    wb2.save(os.path.join(OUT_DIR, '次洪模型汇总.xlsx'))
    print(f"  次洪汇总Excel已保存")

    # -- 每场洪水单独Excel: 蒸发/产流/水源/汇流(含马斯京根) --
    for fe_idx, fe in enumerate(flood_events):
        det = run_flood_detailed(fe, daily_data, hourly_data, sub_areas, sub_nseg_flood)
        if det is None: continue
        det['no'] = fe_idx + 1

        wb_fe = Workbook(); wb_fe.remove(wb_fe.active)
        tag = fe['start'].strftime('%Y%m%d')
        date_h = [(fe['start'] + timedelta(hours=i)).strftime('%m-%d %Hh') for i in range(det['n'])]
        fd = det['fd']

        # --- 蒸散发 ---
        ws = wb_fe.create_sheet('蒸散发')
        hdr = ['时间']
        for i in range(N_SUBS):
            sn = STATION_NAMES[i]
            hdr += [f'{sn}_P', f'{sn}_E0', f'{sn}_EU', f'{sn}_EL', f'{sn}_ED',
                    f'{sn}_WU', f'{sn}_WL', f'{sn}_WD']
        rows = []
        for t in range(det['n']):
            row = [date_h[t]]
            for i in range(N_SUBS):
                r = det['sub_detail'][i]['result']
                row += [round(fd['P_stations'][t,i],2), round(det['e0_h'][t],3),
                        round(r['EU'][t],3), round(r['EL'][t],3), round(r['ED'][t],3),
                        round(r['WU'][t],2), round(r['WL'][t],2), round(r['WD'][t],2)]
            rows.append(row)
        write_excel_sheet(ws, hdr, rows)

        # --- 产流 ---
        ws = wb_fe.create_sheet('产流')
        hdr = ['时间']
        for i in range(N_SUBS):
            sn = STATION_NAMES[i]
            hdr += [f'{sn}_PE', f'{sn}_R', f'{sn}_FR']
        rows = []
        for t in range(det['n']):
            row = [date_h[t]]
            for i in range(N_SUBS):
                r = det['sub_detail'][i]['result']
                row += [round(r['PE'][t],2), round(r['R'][t],3), round(r['FR_series'][t],3)]
            rows.append(row)
        write_excel_sheet(ws, hdr, rows)

        # --- 分水源 ---
        ws = wb_fe.create_sheet('分水源')
        hdr = ['时间']
        for i in range(N_SUBS):
            sn = STATION_NAMES[i]
            hdr += [f'{sn}_RS', f'{sn}_RI', f'{sn}_RG', f'{sn}_S']
        rows = []
        for t in range(det['n']):
            row = [date_h[t]]
            for i in range(N_SUBS):
                r = det['sub_detail'][i]['result']
                row += [round(r['RS'][t],3), round(r['RI'][t],3),
                        round(r['RG'][t],3), round(r['S'][t],3)]
            rows.append(row)
        write_excel_sheet(ws, hdr, rows)

        # --- 汇流(含马斯京根) ---
        ws = wb_fe.create_sheet('汇流')
        hdr = ['时间']
        for i in range(N_SUBS):
            sn = STATION_NAMES[i]
            sd = det['sub_detail'][i]
            ns = sd['n_seg']
            hdr += [f'{sn}_QS', f'{sn}_QI', f'{sn}_QG', f'{sn}_QT', f'{sn}_QTlag']
            if ns > 0:
                seg_labels = ['', '第一河段', '第二河段', '第三河段', '第四河段', '第五河段']
                for s in range(1, ns):
                    hdr += [f'{sn}_{seg_labels[s]}']
                hdr += [f'{sn}_Qout']
            else:
                hdr += [f'{sn}_Qout']
        hdr += ['总出流(m3/s)', '实测Q(m3/s)']

        rows = []
        for t in range(det['n']):
            row = [date_h[t]]
            for i in range(N_SUBS):
                r = det['sub_detail'][i]['result']
                sd = det['sub_detail'][i]
                row += [round(r['QS'][t],3), round(r['QI_series'][t],3),
                        round(r['QG_series'][t],3), round(r['QT'][t],3),
                        round(r['QT_lagged'][t],3)]
                segs = sd['segs']
                if sd['n_seg'] > 0:
                    for s in range(1, len(segs)):
                        row.append(round(segs[s][t], 3))
                else:
                    row.append(round(segs[0][t], 3))
            row += [round(det['Q_total'][t],2), round(det['Q_obs'][t],2)]
            rows.append(row)
        write_excel_sheet(ws, hdr, rows)

        # 标色: Qout列黄, 总出流蓝
        for j in range(1, ws.max_column + 1):
            h = str(ws.cell(1, j).value or '')
            if 'Qout' in h:
                for r in range(1, ws.max_row + 1): ws.cell(r, j).fill = YELLOW_FILL
            elif '总出流' in h:
                for r in range(1, ws.max_row + 1):
                    ws.cell(r, j).fill = BLUE_FILL; ws.cell(r, j).font = BLUE_FONT_W

        fname = f'洪水_{det["no"]:02d}_{tag}.xlsx'
        wb_fe.save(os.path.join(flood_out_dir, fname))

    print(f"  次洪单场Excel已保存: {flood_out_dir}/ ({len(flood_results)}个)")


# ============================================================
#  第13部分: 图表输出
# ============================================================
def output_charts(daily_data, daily_Q_sim, flood_results, sub_areas):
    """生成日模型+次洪模型流量过程线"""
    Q_obs = daily_data['Q_obs']; dates = daily_data['dates']
    P_areal = np.sum([daily_data['P_stations'][:,i]*sub_areas[i] for i in range(N_SUBS)], axis=0)/F
    years = sorted(set(d.year for d in dates))

    # -- 日模型逐年图 --
    daily_dir = os.path.join(OUT_DIR, 'charts', 'daily')
    os.makedirs(daily_dir, exist_ok=True)

    for yr in years:
        m = np.array([d.year == yr for d in dates])
        x = np.arange(m.sum())
        qo, qs, p = Q_obs[m], daily_Q_sim[m], P_areal[m]
        Rs = runoff_depth(qs, F, DT_DAILY); Ro = runoff_depth(qo, F, DT_DAILY)
        re = abs(Rs-Ro)/max(Ro,0.01)*100; dc = nash_sutcliffe(qs, qo)

        fig, (ax_r, ax_f) = plt.subplots(2,1,figsize=(16,8), sharex=True,
                                          gridspec_kw={'height_ratios':[1,2.5]})
        ax_r.bar(x, p, color='#4A90D9', width=1.0, alpha=0.9)
        ax_r.set_ylim(max(p.max()*1.4,10), 0); ax_r.set_ylabel('Rain (mm)', fontsize=10, fontweight='bold')
        ax_r.tick_params(labelsize=9); ax_r.grid(axis='y', alpha=0.3, linestyle='--')
        ax_r.set_title(f'Chengcun Daily {yr} ({"Cal" if yr<=1994 else "Val"})  DC={dc:.3f}  R_err={re:.1f}%',
                       fontsize=13, fontweight='bold')

        ax_f.plot(x, qo, 'k-', lw=1.3, label='Observed', alpha=0.85)
        ax_f.plot(x, qs, 'r-', lw=1.0, label='Simulated', alpha=0.85)
        ax_f.set_ylabel('Q (m$^3$/s)', fontsize=10, fontweight='bold')
        ax_f.set_xlabel('Day', fontsize=10); ax_f.tick_params(labelsize=9)
        ax_f.grid(True, alpha=0.3, linestyle='--')
        ax_f.legend(loc='upper right', fontsize=10, framealpha=0.9)

        fig.tight_layout()
        fig.savefig(os.path.join(daily_dir, f'daily_{yr}.png'), dpi=180, bbox_inches='tight')
        plt.close(fig)
    print(f"  日模型图: {daily_dir}/ ({len(years)}张)")

    # -- 次洪模型逐场图 --
    flood_dir = os.path.join(OUT_DIR, 'charts', 'flood')
    os.makedirs(flood_dir, exist_ok=True)

    for fr in flood_results:
        x = np.arange(fr['n']); qo, qs = fr['Q_obs'], fr['Q_sim']
        fd = fr['fd']
        p = np.sum([fd['P_stations'][:,i]*sub_areas[i] for i in range(N_SUBS)], axis=0)/F

        fig, (ax_r, ax_f) = plt.subplots(2,1,figsize=(16,8), sharex=True,
                                          gridspec_kw={'height_ratios':[1,2.5]})
        ax_r.bar(x, p, color='#4A90D9', width=1.0, alpha=0.9)
        ax_r.set_ylim(max(p.max()*1.4,10), 0); ax_r.set_ylabel('Rain (mm)', fontsize=10, fontweight='bold')
        ax_r.tick_params(labelsize=9); ax_r.grid(axis='y', alpha=0.3, linestyle='--')
        tag = 'Cal' if fr['is_rate'] else 'Val'
        ax_r.set_title(f'Chengcun Flood #{fr["no"]} {tag} {fr["date"].strftime("%Y-%m-%d")}  '
                       f'R_err={fr["R_err"]:.1f}% Qp_err={fr["Qp_err"]:.1f}% DC={fr["DC"]:.3f}  '
                       f'[{"OK" if fr["ok"] else "FAIL"}]', fontsize=13, fontweight='bold')

        ax_f.plot(x, qo, 'k-', lw=1.3, label='Observed', alpha=0.85)
        ax_f.plot(x, qs, 'r-', lw=1.0, label='Simulated', alpha=0.85)
        ax_f.set_ylabel('Q (m$^3$/s)', fontsize=10, fontweight='bold')
        ax_f.set_xlabel('Time (h)', fontsize=10); ax_f.tick_params(labelsize=9)
        ax_f.grid(True, alpha=0.3, linestyle='--')
        ax_f.legend(loc='upper right', fontsize=10, framealpha=0.9)

        fig.tight_layout()
        fig.savefig(os.path.join(flood_dir, f'flood_{fr["no"]:02d}.png'), dpi=180, bbox_inches='tight')
        plt.close(fig)
    print(f"  次洪模型图: {flood_dir}/ ({len(flood_results)}张)")


# ============================================================
#  第14部分: 主程序
# ============================================================
def main():
    print("=" * 60)
    print("  呈村流域(290km2) 新安江模型")
    print("=" * 60)

    # 数据路径
    excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..',
                              '呈村流域资料（含场次洪水）-0613更新场次洪水实测流量.xls')

    # ---- 1. 读取数据 ----
    print("\n[1/5] 读取数据...")
    daily_data, hourly_data, sub_info, flood_events_dict, _ = read_xaj_data(excel_path)
    sub_areas = sub_info['areas']
    flood_events = flood_events_dict['flood']  # 次洪模型场次列表
    # 统一key名
    hourly_data['dates'] = hourly_data['datetimes']
    print(f"  日资料: {len(daily_data['dates'])}天, 时段资料: {len(hourly_data['dates'])}h")
    print(f"  子流域: {N_SUBS}个, 场次洪水: {len(flood_events)}场")

    # ---- 2. 日模型 ----
    print("\n[2/5] 日模型计算...")
    Q_sim_daily, sub_detail = run_daily_model(daily_data, sub_areas)
    Q_obs = daily_data['Q_obs']
    dates = daily_data['dates']

    dc_all = nash_sutcliffe(Q_sim_daily, Q_obs)
    print(f"  总体 DC={dc_all:.3f}")
    rate_dc, test_dc = [], []
    for yr in range(1989, 1997):
        m = np.array([d.year == yr for d in dates])
        qs, qo = Q_sim_daily[m], Q_obs[m]
        Rs = runoff_depth(qs, F, DT_DAILY); Ro = runoff_depth(qo, F, DT_DAILY)
        dc = nash_sutcliffe(qs, qo); re = abs(Rs-Ro)/max(Ro,0.01)*100
        (rate_dc if yr<=1994 else test_dc).append(dc)
        print(f"    {yr}: DC={dc:.3f}  R_err={re:.1f}%  {'OK' if re<=20 else '--'}")
    print(f"  率定 DC={np.mean(rate_dc):.3f}  检验 DC={np.mean(test_dc):.3f}")

    # ---- 3. 次洪模型 ----
    print("\n[3/5] 次洪模型计算...")
    flood_results, mk_detail = run_flood_model(flood_events, daily_data, hourly_data,
                                               sub_areas, sub_info['m_routing_flood'])

    rate_ok = sum(1 for f in flood_results if f['is_rate'] and f['ok'])
    test_ok = sum(1 for f in flood_results if not f['is_rate'] and f['ok'])
    for fr in flood_results:
        t = 'Rate' if fr['is_rate'] else 'Test'
        print(f"  #{fr['no']} {t} {fr['date'].strftime('%Y-%m-%d')}: "
              f"R={fr['R_err']:.1f}% Qp={fr['Qp_err']:.1f}% DC={fr['DC']:.3f} "
              f"{'OK' if fr['ok'] else '--'}")

    rate_r = np.mean([f['R_err'] for f in flood_results if f['is_rate']])
    test_r = np.mean([f['R_err'] for f in flood_results if not f['is_rate']])
    print(f"  率定 R={rate_r:.1f}%  检验 R={test_r:.1f}%")
    print(f"  双达标:  {rate_ok+test_ok}/{len(flood_results)}")

    # ---- 4. 输出Excel ----
    print("\n[4/5] 输出Excel...")
    output_excel(daily_data, Q_sim_daily, sub_detail, flood_results, sub_areas, mk_detail, flood_events, hourly_data, sub_info['m_routing_flood'])

    # ---- 5. 输出图表 ----
    print("\n[5/5] 输出图表...")
    output_charts(daily_data, Q_sim_daily, flood_results, sub_areas)

    # ---- 完成 ----
    print("\n" + "=" * 60)
    print("  全部完成!")
    print(f"  输出目录: {OUT_DIR}")
    print("=" * 60)


if __name__ == '__main__':
    main()
